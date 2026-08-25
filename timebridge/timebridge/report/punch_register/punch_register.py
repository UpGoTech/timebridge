# Copyright (c) 2026, UPGO and contributors
# For license information, please see license.txt

"""
Everyone's clock times for a month, laid out sideways.

The Attendance Register answers "who was here" in one letter a day. TimeBridge Employee
Attendance Detail answers "what time" — but for one person, reading downwards,
which means sixteen exports to see a team.

This is that same question asked sideways: staff down the page, dates across
it, and the actual In and Out in every cell. It exists for Excel. Thirty-one
date columns will not fit a sheet of A4, and the letter register is already
the thing that prints.
"""

import calendar
import io

import frappe

from frappe.utils import cint, get_last_day, getdate, now_datetime

from timebridge.timebridge.report.attendance_report.attendance_report import (
    SICK_LEAVE_CODE,
    STATUS_CODE,
    WEEKLY_OFF_CODE,
    code_sort_key,
    get_employees,
    weekly_off_dates,
)

# Both times in one cell, as chosen. The alternative — a separate In column and
# Out column per day — reaches sixty-four columns and puts each time in its own
# Excel cell, which is better for arithmetic and worse for reading. Reading won.
TIME_JOIN = "-"

# Worded like the register's own legend, and in English for the same reason:
# every other label in this application is, and a sheet that switches language
# halfway reads as unfinished.
#
# Kept as separate items so the workbook can print the same legend without the
# markup, rather than the two drifting apart as one string is edited.
LEGEND_ITEMS = (
    "Each cell shows <b>In-Out</b> time",
    "11:20- = no Out punch",
    "R = Rest (Weekly Off)",
    "H = Holiday",
    "S = Sick Leave",
    "L = Other Leave",
    "A = Absent",
    "Blank = no record",
)

LEGEND = " &nbsp; ".join(LEGEND_ITEMS)


def execute(filters=None):

    filters = frappe._dict(filters or {})

    month, year = resolve_period(filters)
    days = calendar.monthrange(year, month)[1]

    off_days = weekly_off_dates(year, month, days)
    employees = get_employees(filters)

    columns = build_columns(days, off_days)
    heading = print_heading(filters, month, year)

    if not employees:
        return columns, [], heading

    records = day_times(employees, year, month)

    attach_machine_user_ids(employees, filters.get("biometric_machine"))

    # By the id on this terminal, read as a number — 2 before 10, not after it.
    employees = sorted(employees, key=punch_code_sort_key)

    return columns, build_rows(employees, records, off_days, days), heading


def resolve_period(filters):

    if filters.get("month") and filters.get("year"):
        return cint(filters.month), cint(filters.year)

    today = getdate()

    return today.month, today.year


# ---------------------------------------------------------------- heading

def print_heading(filters, month, year):
    """
    The title block above the table.

    Hidden on screen and shown when printed, the same way the register does it:
    on screen the month and the machine are already sitting in the filter bar a
    few pixels above, and repeating them only pushes the table further down.
    The legend stays visible in both, since the cells mean nothing without it.

    The Excel export does not use this. Frappe writes only the grid into a
    workbook, so that file builds its own title rows — see `export_excel`.
    """

    lines = [
        (organisation_name(filters), "font-size:15px;font-weight:700"),
        (f"Punch Register for {period_name(month, year)}", "font-weight:600"),
        (filter_line(filters), "font-size:11px;color:#777"),
    ]

    block = "".join(
        f"<div style='{style}'>{frappe.utils.escape_html(text)}</div>"
        for text, style in lines
        if text
    )

    return (
        "<style>"
        "  .tb-print-title { display: none; }"
        "  @media print { .tb-print-title { display: block !important; } }"
        "</style>"
        "<div class='tb-print-title' style='text-align:center;line-height:1.5;"
        f"margin-bottom:6px'>{block}</div>"
        f"<div style='text-align:center;font-size:11px;color:#777'>{LEGEND}</div>"
    )


def attach_machine_user_ids(employees, machine_id):
    """
    Stamp each row with the id this terminal actually enrolled.

    TimeBridge Employee Code is unique across every machine, so a second terminal's user
    `4` becomes `AIFACE002-4` and then looks missing when the sheet is read as
    1, 2, 3, 6. The device id does not collide: it is unique per machine, which
    is the only id this report is answering for once a machine is chosen.
    """

    if not machine_id:
        return

    by_employee = {}

    for row in frappe.get_all(
        "TimeBridge Machine User",
        filters={
            "machine": machine_id,
            "employee": ["in", [emp.name for emp in employees]],
        },
        fields=["employee", "user_id"],
    ):
        by_employee.setdefault(row.employee, []).append(row.user_id)

    for emp in employees:

        ids = sorted(
            by_employee.get(emp.name) or [],
            key=lambda value: code_sort_key(frappe._dict(employee_code=value)),
        )

        # Two enrolments on one person (09 and F09) stay visible rather than
        # picking a winner — the punches on the row already belong to both.
        emp.machine_user_id = ", ".join(ids) if ids else None


def punch_code_sort_key(employee):

    shown = (employee.get("machine_user_id") or employee.employee_code or "")
    first = shown.split(",")[0].strip()

    return code_sort_key(frappe._dict(employee_code=first))


def organisation_name(filters):
    """
    Whose staff this is, or nothing if the sheet cannot say honestly.

    With no TimeBridge Organization chosen the rows may span several, so the name is only
    filled in when there is exactly one on the site — naming the wrong company
    at the top of a signed sheet is worse than leaving it off.
    """

    if filters.get("organization"):
        return link_label("TimeBridge Organization", filters.get("organization"))

    names = frappe.get_all("TimeBridge Organization", pluck="name", limit=2)

    return link_label("TimeBridge Organization", names[0]) if len(names) == 1 else None


def period_name(month, year):

    return f"{calendar.month_name[month]} {year}"


def link_label(doctype, value):
    """
    The name a person would recognise rather than the id.

    Every doctype here is named by a code — SH-00001 in a heading tells the
    reader nothing about which shift they are looking at.
    """

    if not value:
        return None

    title_field = frappe.get_meta(doctype).get_title_field()

    return frappe.db.get_value(doctype, value, title_field) or value


def machine_label(machine):
    """
    How a terminal is named on the sheet: "Fabrixcel (BM-106762)".

    An unset filter is spelled out rather than left blank, because a sheet
    holding every machine and saying nothing looks exactly like a sheet from
    one machine — which is the confusion this heading exists to end.
    """

    if not machine:
        return "All Machines"

    name = link_label("TimeBridge Machine", machine)

    return machine if name == machine else f"{name} ({machine})"


def filter_line(filters):
    """
    What was narrowed, in the order somebody would say it.

    Only the filters that were set — a line listing every empty one is noise.
    The machine is the exception and always appears, per `machine_label`.
    """

    parts = [f"Machine: {machine_label(filters.get('biometric_machine'))}"]

    for label, fieldname, doctype in (
        ("TimeBridge Branch", "branch", "TimeBridge Branch"),
        ("TimeBridge Department", "department", "TimeBridge Department"),
        ("TimeBridge Shift", "shift", "TimeBridge Shift"),
        ("TimeBridge Employee", "employee", "TimeBridge Employee"),
    ):
        value = link_label(doctype, filters.get(fieldname))

        if value:
            parts.append(f"{label}: {value}")

    if filters.get("include_inactive"):
        parts.append("Including inactive employees")

    return "     ".join(parts)


def day_times(employees, year, month):
    """
    Clock times for the month, keyed by (employee, day-of-month).

    The register has a helper of its own, but it never selects the times —
    it only ever needed a status letter. Rather than widening a query two
    reports would then share for different reasons, this asks for exactly what
    a punch register needs, and needs no join: leave_type is already carried on
    the attendance row.
    """

    first = getdate(f"{year}-{month:02d}-01")
    last = get_last_day(first)

    rows = frappe.db.sql(
        """
        SELECT employee, DAY(attendance_date) AS day,
               first_in, last_out, status, leave_type
        FROM `tabTimeBridge Attendance`
        WHERE employee IN %(employees)s
          AND attendance_date BETWEEN %(first)s AND %(last)s
        """,
        {"employees": [e.name for e in employees], "first": first, "last": last},
        as_dict=True,
    )

    return {(r.employee, r.day): r for r in rows}


def build_columns(days, off_days):

    columns = [
        {"label": "Code", "fieldname": "employee_code", "fieldtype": "Data", "width": 60},
        {"label": "Full Name Of The TimeBridge Employee", "fieldname": "employee_name",
         "fieldtype": "Data", "width": 180},
    ]

    for day in range(1, days + 1):

        columns.append({
            "label": str(day),
            "fieldname": f"day_{day}",
            "fieldtype": "Data",
            # Wide enough for "11:38-19:01" without clipping, which is the
            # whole point of the report.
            "width": 100,
            # Read by the client to shade the date of a weekly off, the same
            # way the register marks its Sundays.
            "weekly_off": 1 if day in off_days else 0,
        })

    return columns


def build_rows(employees, records, off_days, days):

    data = []

    for emp in employees:

        row = {
            "employee": emp.name,
            "employee_code": emp.get("machine_user_id") or emp.employee_code,
            "employee_name": emp.employee_name,
        }

        for day in range(1, days + 1):
            row[f"day_{day}"] = cell(records.get((emp.name, day)), day, off_days)

        data.append(row)

    return data


def cell(record, day, off_days):
    """
    What one day says.

    A time when there is one. Otherwise the register's own letter, so a day off
    and a day skipped never look alike — an empty cell would make a Sunday
    indistinguishable from an absence.
    """

    if not record:
        return WEEKLY_OFF_CODE if day in off_days else ""

    first_in = clock(record.first_in)
    last_out = clock(record.last_out)

    if first_in and last_out:
        return f"{first_in}{TIME_JOIN}{last_out}"

    # One punch and no pair. The time is still real and worth showing; the
    # trailing dash says the other half never arrived rather than implying the
    # person left at that moment.
    if first_in:
        return f"{first_in}{TIME_JOIN}"

    if record.status == "On Leave" and (record.leave_type or "") == "Sick Leave":
        return SICK_LEAVE_CODE

    return STATUS_CODE.get(record.status, "") or (
        WEEKLY_OFF_CODE if day in off_days else ""
    )


def clock(value):
    """Just the clock face — the date is already the column heading."""

    return str(value)[11:16] if value else None


# ------------------------------------------------------------------ excel
#
# openpyxl is imported inside the functions below rather than at the top of the
# module. This module is loaded every time the report is drawn on screen, and
# none of it is needed until somebody asks for the file.

# Excel measures a column in characters, not pixels, so these are unrelated to
# the widths the columns carry for the screen. A day column has to hold
# "11:38-19:01" — eleven characters — without spilling into its neighbour,
# which is exactly what Frappe's own export got wrong: it divides the declared
# width by ten, leaving every time column too narrow to show a time.
SERIAL_WIDTH = 5
COLUMN_WIDTHS = {"employee_code": 11, "employee_name": 32}
DAY_WIDTH = 12.5

# Serial number, code and name are held in place while the dates scroll, so a
# reader thirty columns across still knows whose row they are on. The third of
# them is the name, which is also the only column read as words rather than
# centred under its heading.
FROZEN_COLUMNS = 3

MUTED = "FF6C757D"
GRID = "FFBFBFBF"
HEADER_BG = "FFE9ECEF"

# A weekly off is tinted the same red the register uses on screen, and the
# column beneath it far more faintly — over thirty-one columns the header alone
# is too far away to place a date by the time the eye is at the bottom.
OFF_HEADER_BG = "FFFFD7D7"
OFF_BODY_BG = "FFFDF2F2"

# The letters the screen colours, in the nearest fixed hex: a workbook has no
# access to the theme variables punch_register.js uses in the browser.
LETTER_COLOUR = {
    WEEKLY_OFF_CODE: "FF8D96A0",
    "H": "FF7C3AED",
    SICK_LEAVE_CODE: "FF2563EB",
    "L": "FF3B82F6",
    "A": "FFDC2626",
    "?": "FFF59E0B",
}


@frappe.whitelist()
def export_excel(filters=None):
    """
    The workbook this report is actually for.

    Frappe's own export writes the grid and nothing else: no title, no machine,
    no month, no legend, no frozen header, and column widths a tenth of the
    declared ones. On a sheet of thirty-one time columns that produced a wall
    of numbers running into each other with nothing to say where they came
    from. None of it is reachable from a column definition, so the file is
    built here instead.

    The response is the file itself, which is why the client posts a form at
    this rather than calling it — see `download_excel` in punch_register.js.
    """

    from frappe.core.doctype.access_log.access_log import make_access_log
    from frappe.desk.utils import provide_binary_file
    from frappe.permissions import can_export

    # The same gate Frappe's own export passes through, on the same doctype
    # this report is declared against.
    can_export("TimeBridge Attendance", raise_exception=True)

    filters = frappe._dict(frappe.parse_json(filters) or {})

    columns, data = execute(filters)[:2]

    make_access_log(
        doctype="TimeBridge Attendance",
        report_name="Punch Register",
        file_type="Excel",
        method="Export",
        filters=filters,
    )

    month, year = resolve_period(filters)

    provide_binary_file(
        f"Punch Register {period_name(month, year)}",
        "xlsx",
        build_workbook(columns, data, filters, month, year),
    )


def build_workbook(columns, data, filters, month, year):

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Punch Register"

    # The serial number is added here rather than in the report: on screen
    # Frappe numbers the rows itself, and a second column of counting beside
    # its own would look like a mistake.
    headers = ["Sr"] + [column["label"] for column in columns]
    widths = [SERIAL_WIDTH] + [column_width(column) for column in columns]
    weekly_off = [False] + [bool(column.get("weekly_off")) for column in columns]

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    last_column = len(headers)

    header_row = write_title(sheet, filters, month, year, last_column)
    write_header(sheet, header_row, headers, weekly_off)

    for number, record in enumerate(data, start=1):
        write_row(sheet, header_row + number, number, record, columns, weekly_off)

    last_row = header_row + len(data)

    if not data:
        text_line(sheet, last_row + 1, "No attendance records for this period.",
                  last_column, size=10, italic=True, colour=MUTED)

    set_layout(sheet, header_row, last_row)

    stream = io.BytesIO()
    workbook.save(stream)

    return stream.getvalue()


def column_width(column):

    return COLUMN_WIDTHS.get(column["fieldname"], DAY_WIDTH)


def write_title(sheet, filters, month, year, last_column):
    """
    The block above the table, and the reason this export exists: a page of
    times that does not say whose staff, which terminal or which month is not
    a document somebody can file.

    Returns the row the header belongs on, one blank row further down.
    """

    row = 1

    organisation = organisation_name(filters)

    if organisation:
        text_line(sheet, row, organisation, last_column,
                  size=14, bold=True, height=22)
        row += 1

    text_line(sheet, row, f"Punch Register — {period_name(month, year)}",
              last_column, size=12, bold=True, height=18)
    row += 1

    text_line(sheet, row, filter_line(filters), last_column, colour=MUTED)
    row += 1

    text_line(sheet, row, f"Generated {now_datetime().strftime('%d-%m-%Y %H:%M')}",
              last_column, size=9, italic=True, colour=MUTED)

    return row + 2


def text_line(sheet, row, text, last_column=None, size=10, bold=False,
              italic=False, colour=None, height=None):
    """
    One line of prose above or below the table.

    Title rows are merged across the sheet and centred. A merged range that
    also crosses a frozen *column* split is clipped at the split, so the
    names are not frozen — only the header row is, which these lines sit
    above and never cross.
    """

    from openpyxl.styles import Alignment, Font

    if last_column:
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column
        )

    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(size=size, bold=bold, italic=italic, color=colour)
    cell.alignment = Alignment(
        horizontal="center" if last_column else "left",
        vertical="center",
    )

    if height:
        sheet.row_dimensions[row].height = height


def write_header(sheet, row, headers, weekly_off):

    from openpyxl.styles import Alignment, Font, PatternFill

    sheet.row_dimensions[row].height = 20

    for index, label in enumerate(headers, start=1):

        cell = sheet.cell(row=row, column=index, value=label)
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = grid_border()
        cell.fill = PatternFill(
            "solid", fgColor=OFF_HEADER_BG if weekly_off[index - 1] else HEADER_BG
        )


def write_row(sheet, row, number, record, columns, weekly_off):

    from openpyxl.styles import Alignment, Font, PatternFill

    values = [number] + [record.get(column["fieldname"]) or "" for column in columns]

    for index, value in enumerate(values, start=1):

        cell = sheet.cell(row=row, column=index, value=value)
        cell.border = grid_border()

        colour = LETTER_COLOUR.get(value)

        # Absence is the one worth spotting without reading the cell, the same
        # way the screen weights it.
        cell.font = Font(size=10, color=colour, bold=value == "A")

        # Only the name is read as words; everything else is a code or a time
        # and centres under its heading.
        cell.alignment = Alignment(
            horizontal="left" if index == FROZEN_COLUMNS else "center",
            vertical="center",
        )

        if weekly_off[index - 1]:
            cell.fill = PatternFill("solid", fgColor=OFF_BODY_BG)


def grid_border():

    from openpyxl.styles import Border, Side

    line = Side(style="thin", color=GRID)

    return Border(left=line, right=line, top=line, bottom=line)


def set_layout(sheet, header_row, last_row):
    """
    What makes thirty-four columns readable once they are in Excel.

    Only the header row is frozen. The names used to freeze too, but a
    column freeze splits every merged title at D and cuts the centred
    heading in half — which is the thing this file is for.
    """

    from openpyxl.worksheet.properties import PageSetupProperties

    # Header row only. Freezing the name columns as well would split every
    # merged title at column D and cut the centred heading in half.
    sheet.freeze_panes = f"A{header_row + 1}"

    if last_row > header_row:
        # Over the name and code only. Sorting or hiding by a single day is
        # never the question being asked, and thirty-one extra arrows across
        # the header would bury the two that are.
        sheet.auto_filter.ref = f"A{header_row}:C{last_row}"

    # Repeat the dates at the top of every printed page.
    sheet.print_title_rows = f"{header_row}:{header_row}"

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
